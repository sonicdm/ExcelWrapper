from excelwrapper.worksheet.worksheet import Worksheet
from excelwrapper.Workbook.workbook import Workbook


class Xlsx:
    def __init__(self):
        pass

    def read(self, f, **kwargs):
        from openpyxl import load_workbook
        wbin = load_workbook(
            f,
            data_only=True,
            guess_types=True
        )
        wbout = Workbook()
        sheets = {k: v for v, k in enumerate(wbin.worksheets)}
        workbook = []
        for ws in wbin.worksheets:
            newsheet = Worksheet(ws.title)
            rowlist = []
            for row in ws.rows:
                for cell in row:
                    pass
            workbook.append(rowlist)
        return dict(sheetindex=sheets, workbook=workbook)

    def write(self, i, j):
        pass
